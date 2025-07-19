from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
from datetime import datetime
import uuid
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io

app = Flask(__name__)
app.secret_key = 'barakah_secret_key_2024'

# Configuration
UPLOAD_FOLDER = 'uploads'
EXCEL_FILE = 'data_pendaftar.xlsx'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('static/images', exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_excel():
    """Initialize Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pendaftar"
        
        # Headers
        headers = [
            'ID', 'Tanggal Daftar', 'Nama Lengkap', 'NIK', 'Nomor KK', 'NISN', 
            'Tempat Lahir', 'Tanggal Lahir', 'Jenis Kelamin', 'Agama', 
            'Alamat Jalan', 'RT', 'RW', 'Dusun', 'Desa', 'Kab/Kota', 'Provinsi',
            'Diterima Sebagai', 'Diterima di Kelas', 'No WA',
            'Jenis Tinggal', 'Alat Transportasi', 'No KIP/KKS/PKH', 'Anak Ke', 
            'Jumlah Saudara', 'Berat Badan', 'Tinggi Badan', 'Penyakit',
            'Lulusan Dari', 'Tanggal Lulus', 'No Ijazah', 'Asal Sekolah', 
            'Tahun Keluar', 'Hobi', 'Cita-cita', 'Pekerjaan',
            'Nama Ayah', 'NIK Ayah', 'TTL Ayah', 'Agama Ayah', 'Pendidikan Ayah', 
            'Pekerjaan Ayah', 'Penghasilan Ayah',
            'Nama Ibu', 'NIK Ibu', 'TTL Ibu', 'Agama Ibu', 'Pendidikan Ibu', 
            'Pekerjaan Ibu', 'Penghasilan Ibu',
            'Moda Transportasi', 'Jarak ke Sekolah',
            'File Foto', 'File KK', 'File Ijazah', 'File Akta',
            'Perkembangan Warga Belajar'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_registration():
    try:
        # Generate unique ID
        registration_id = str(uuid.uuid4())[:8]
        
        # Handle file uploads
        uploaded_files = {}
        file_fields = ['foto', 'kk', 'ijazah', 'akta']
        
        for field in file_fields:
            if field in request.files:
                file = request.files[field]
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(f"{registration_id}_{field}_{file.filename}")
                    file_path = os.path.join(UPLOAD_FOLDER, filename)
                    file.save(file_path)
                    uploaded_files[field] = filename
                else:
                    uploaded_files[field] = ''
            else:
                uploaded_files[field] = ''
        
        # Prepare data for Excel
        data = [
            registration_id,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            request.form.get('nama_lengkap', ''),
            request.form.get('nik', ''),
            request.form.get('nomor_kk', ''),
            request.form.get('nisn', ''),
            request.form.get('tempat_lahir', ''),
            request.form.get('tanggal_lahir', ''),
            request.form.get('jenis_kelamin', ''),
            request.form.get('agama', ''),
            request.form.get('alamat_jalan', ''),
            request.form.get('rt', ''),
            request.form.get('rw', ''),
            request.form.get('dusun', ''),
            request.form.get('desa', ''),
            request.form.get('kab_kota', ''),
            request.form.get('provinsi', ''),
            request.form.get('diterima_sebagai', ''),
            request.form.get('diterima_kelas', ''),
            request.form.get('no_wa', ''),
            request.form.get('jenis_tinggal', ''),
            request.form.get('alat_transportasi', ''),
            request.form.get('no_kip', ''),
            request.form.get('anak_ke', ''),
            request.form.get('jumlah_saudara', ''),
            request.form.get('berat_badan', ''),
            request.form.get('tinggi_badan', ''),
            request.form.get('penyakit', ''),
            request.form.get('lulusan_dari', ''),
            request.form.get('tanggal_lulus', ''),
            request.form.get('no_ijazah', ''),
            request.form.get('asal_sekolah', ''),
            request.form.get('tahun_keluar', ''),
            request.form.get('hobi', ''),
            request.form.get('cita_cita', ''),
            request.form.get('pekerjaan', ''),
            request.form.get('nama_ayah', ''),
            request.form.get('nik_ayah', ''),
            request.form.get('ttl_ayah', ''),
            request.form.get('agama_ayah', ''),
            request.form.get('pendidikan_ayah', ''),
            request.form.get('pekerjaan_ayah', ''),
            request.form.get('penghasilan_ayah', ''),
            request.form.get('nama_ibu', ''),
            request.form.get('nik_ibu', ''),
            request.form.get('ttl_ibu', ''),
            request.form.get('agama_ibu', ''),
            request.form.get('pendidikan_ibu', ''),
            request.form.get('pekerjaan_ibu', ''),
            request.form.get('penghasilan_ibu', ''),
            request.form.get('moda_transportasi', ''),
            request.form.get('jarak_sekolah', ''),
            uploaded_files.get('foto', ''),
            uploaded_files.get('kk', ''),
            uploaded_files.get('ijazah', ''),
            uploaded_files.get('akta', ''),
            ''  # Perkembangan Warga Belajar (empty for new registrations)
        ]
        
        # Save to Excel
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(data)
        wb.save(EXCEL_FILE)
        
        flash('Pendaftaran berhasil! Silakan bergabung dengan grup WhatsApp.', 'success')
        return redirect(url_for('success'))
        
    except Exception as e:
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/success')
def success():
    return render_template('success.html')

@app.route('/admin')
def admin_login():
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_post():
    username = request.form.get('username')
    password = request.form.get('password')
    
    if username == 'admin' and password == 'barakah123':
        session['admin_logged_in'] = True
        return redirect(url_for('admin_dashboard'))
    else:
        flash('Username atau password salah!', 'error')
        return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Read data from Excel
    data = []
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                data.append(dict(zip(headers, row)))
    
    return render_template('admin_dashboard.html', data=data)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/download_excel')
def download_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name=f'data_pendaftar_{datetime.now().strftime("%Y%m%d")}.xlsx')
    else:
        flash('File Excel tidak ditemukan!', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/export_formulir/<student_id>')
def export_formulir(student_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Get student data
    student_data = get_student_data(student_id)
    if not student_data:
        flash('Data siswa tidak ditemukan!', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    story.append(Paragraph("FORMULIR PENDAFTARAN MURID BARU", title_style))
    story.append(Paragraph("PKBM AL BARAKAH NPSN : P2952809", title_style))
    story.append(Paragraph("Kecamatan Jiwan Kabupaten Madiun", title_style))
    story.append(Spacer(1, 20))
    
    # Create form data table
    form_data = [
        ['IDENTITAS CALON PESERTA DIDIK', ''],
        ['Nama', student_data.get('Nama Lengkap', '')],
        ['NIS', ''],
        ['NISN', student_data.get('NISN', '')],
        ['Nama Lengkap', student_data.get('Nama Lengkap', '')],
        ['Nomor Induk Kependudukan', student_data.get('NIK', '')],
        ['Jenis Kelamin', student_data.get('Jenis Kelamin', '')],
        ['Kelas', student_data.get('Diterima di Kelas', '')],
        ['Tempat, Tanggal Lahir', f"{student_data.get('Tempat Lahir', '')}, {student_data.get('Tanggal Lahir', '')}"],
        ['Tinggi / Berat Badan', f"{student_data.get('Tinggi Badan', '')} / {student_data.get('Berat Badan', '')}"],
        ['Agama', student_data.get('Agama', '')],
        ['Asal Sekolah', student_data.get('Asal Sekolah', '')],
        ['Alamat', f"{student_data.get('Alamat Jalan', '')}, RT {student_data.get('RT', '')}, RW {student_data.get('RW', '')}, {student_data.get('Desa', '')}, {student_data.get('Kab/Kota', '')}"],
        ['Tinggal bersama', student_data.get('Jenis Tinggal', '')],
        ['No. Telp/Whatsapp', student_data.get('No WA', '')],
        ['Jumlah Saudara', student_data.get('Jumlah Saudara', '')],
        ['Anak Ke', student_data.get('Anak Ke', '')],
        ['Diterima Sebagai', student_data.get('Diterima Sebagai', '')],
        ['IDENTITAS ORANG TUA', ''],
        ['Nama Ayah', student_data.get('Nama Ayah', '')],
        ['Nomor Induk Kependudukan', student_data.get('NIK Ayah', '')],
        ['Tempat, Tanggal Lahir', student_data.get('TTL Ayah', '')],
        ['Pendidikan Terakhir', student_data.get('Pendidikan Ayah', '')],
        ['Pekerjaan', student_data.get('Pekerjaan Ayah', '')],
        ['Nama Ibu', student_data.get('Nama Ibu', '')],
        ['Nomor Induk Kependudukan', student_data.get('NIK Ibu', '')],
        ['Tempat, Tanggal Lahir', student_data.get('TTL Ibu', '')],
        ['Pendidikan Terakhir', student_data.get('Pendidikan Ibu', '')],
        ['Pekerjaan', student_data.get('Pekerjaan Ibu', '')],
        ['Alamat', f"{student_data.get('Alamat Jalan', '')}, {student_data.get('Desa', '')}, {student_data.get('Kab/Kota', '')}"]
    ]
    
    table = Table(form_data, colWidths=[3*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('BACKGROUND', (0, 18), (-1, 18), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'formulir_{student_data.get("Nama Lengkap", "siswa")}.pdf',
        mimetype='application/pdf'
    )

@app.route('/admin/export_buku_induk/<student_id>')
def export_buku_induk(student_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Get student data
    student_data = get_student_data(student_id)
    if not student_data:
        flash('Data siswa tidak ditemukan!', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20,
        alignment=1  # Center
    )
    
    story.append(Paragraph("LEMBAR BUKU INDUK WARGA BELAJAR PAKET B", title_style))
    story.append(Paragraph("NOMOR INDUK/NISN: 5/6", title_style))
    story.append(Spacer(1, 20))
    
    # Create buku induk data
    buku_induk_data = [
        ['A. KETERANGAN TENTANG DIRI WARGA BELAJAR', '', 'Pas Foto 3 x 4'],
        ['Nama Lengkap', ':', student_data.get('Nama Lengkap', '')],
        ['Nomor Induk Kependudukan', ':', student_data.get('NIK', '')],
        ['Tempat, Tanggal Lahir', ':', f"{student_data.get('Tempat Lahir', '')}, {student_data.get('Tanggal Lahir', '')}"],
        ['Jenis Kelamin', ':', student_data.get('Jenis Kelamin', '')],
        ['Agama', ':', student_data.get('Agama', '')],
        ['Pekerjaan', ':', student_data.get('Pekerjaan', '')],
        ['Alamat', ':', f"{student_data.get('Alamat Jalan', '')}, {student_data.get('Desa', '')}, {student_data.get('Kab/Kota', '')}"],
        ['', '', 'Identitas saat diterima'],
        ['No. Telp/Whatsapp', ':', student_data.get('No WA', '')],
        ['B. KETERANGAN PENDIDIKAN WARGA BELAJAR', '', ''],
        ['Pendidikan Sebelumnya', '', ''],
        ['a. Lulusan dari', ':', student_data.get('Lulusan Dari', '')],
        ['b. Tanggal, Bulan, Tahun', ':', student_data.get('Tanggal Lulus', '')],
        ['c. Nomor STTB/Ijazah', ':', student_data.get('No Ijazah', '')],
        ['Mutasi/Pindahan', '', ''],
        ['a. Nama Sekolah', ':', student_data.get('Asal Sekolah', '')],
        ['b. Tahun Keluar', ':', student_data.get('Tahun Keluar', '')],
        ['Diterima di Kelas/Kursus', ':', student_data.get('Diterima di Kelas', '')],
        ['a. Tanggal, Bulan, Tahun', ':', ''],
        ['b. Program', ':', ''],
        ['c. Kelas', ':', student_data.get('Diterima di Kelas', '')],
        ['C. KETERANGAN TENTANG ORANG TUA/WALI', '', ''],
        ['Orang Tua', '', ''],
        ['a. Nama Ayah', ':', student_data.get('Nama Ayah', '')],
        ['   Nama Ibu', ':', student_data.get('Nama Ibu', '')],
        ['b. Agama Ayah', ':', student_data.get('Agama Ayah', '')],
        ['   Agama Ibu', ':', student_data.get('Agama Ibu', '')],
        ['c. Pendidikan Ayah', ':', student_data.get('Pendidikan Ayah', '')],
        ['   Pendidikan Ibu', ':', student_data.get('Pendidikan Ibu', '')],
        ['d. Pekerjaan Ayah', ':', student_data.get('Pekerjaan Ayah', '')],
        ['   Pekerjaan Ibu', ':', student_data.get('Pekerjaan Ibu', '')],
        ['Wali Siswa (Jika Ada)', '', ''],
        ['a. Nama', ':', ''],
        ['b. Agama', ':', ''],
        ['c. Pendidikan', ':', ''],
        ['d. Pekerjaan', ':', ''],
        ['e. Alamat', ':', ''],
        ['', '', ''],
        ['D. KETERANGAN PERKEMBANGAN WARGA BELAJAR', '', ''],
        ['Meninggalkan Kelompok Belajar', '', 'Pas Foto 3 x 4'],
        ['a. Tanggal Meninggalkan', ':', ''],
        ['   Belajar', '', ''],
        ['b. Alasan', ':', ''],
        ['Melanjutkan Ke Sekolah', '', ''],
        ['a. Tanggal Tamat Belajar', ':', ''],
        ['b. Nomor STTB/Ijazah', ':', ''],
        ['', '', 'Identitas saat diterima']
    ]
    
    table = Table(buku_induk_data, colWidths=[2.5*inch, 0.3*inch, 3.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (2, 0), colors.lightgrey),
        ('BACKGROUND', (0, 10), (2, 10), colors.lightgrey),
        ('BACKGROUND', (0, 22), (2, 22), colors.lightgrey),
        ('BACKGROUND', (0, 37), (2, 37), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'buku_induk_{student_data.get("Nama Lengkap", "siswa")}.pdf',
        mimetype='application/pdf'
    )

def get_student_data(student_id):
    """Get student data by ID from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return None
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == student_id:  # ID is in first column
            return dict(zip(headers, row))
    
    return None

if __name__ == '__main__':
    init_excel()
    app.run(debug=True, port=8000)
